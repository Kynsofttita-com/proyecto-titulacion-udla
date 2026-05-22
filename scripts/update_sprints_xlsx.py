"""
Actualiza SPRINTS_PLAN.xlsx con el nuevo plan vertical por grupos (Sprints 5-12).
- Marca el plan original como histórico
- Crea hoja nueva "Sprints 5-12 Vigente" con el plan replanteado
- Actualiza la Overview
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook("SPRINTS_PLAN.xlsx")

# === 1) OVERVIEW ===
ws = wb["Overview"]

ws.cell(row=2, column=1,
        value="Enfoque: HIBRIDO - Sprints 1-4 horizontal (ejecutados). "
              "Sprints 5-12 VERTICAL POR GRUPOS (cambio 2026-05-22, ver PLAN_FASES.md)")

ws.cell(row=5, column=1, value="MICROSERVICIOS (8):")
ws.cell(row=12, column=1, value="  - MS-Reportes")

# Insertar MS-Notificaciones como nueva fila 13
ws.insert_rows(13)
ws.cell(row=13, column=1, value="  - MS-Notificaciones (Emails + in-app)")

# Tras el insert, las filas se corrieron en 1. ROADMAP estaba en 14, ahora 15.
ws.cell(row=15, column=1,
        value="ROADMAP ORIGINAL (Sprints 5-12 SUSTITUIDOS - ver PLAN_FASES.md y hoja "
              "'Sprints 5-12 Vigente'):")

# Marcar Sprints 5-12 como sustituidos (filas 20 a 27 tras el insert)
for sprint_num, row_idx in [(5, 20), (6, 21), (7, 22), (8, 23),
                             (9, 24), (10, 25), (11, 26), (12, 27)]:
    current = ws.cell(row=row_idx, column=1).value or ""
    ws.cell(row=row_idx, column=1, value="  [SUSTITUIDO] " + current.strip())

# Bloque PRINCIPIOS (tras el insert estaba en 28, ahora 29)
ws.cell(row=29, column=1,
        value="PRINCIPIOS DEL DESARROLLO (sprints 1-4 horizontal; sprints 5-12 vertical por grupos):")
ws.cell(row=30, column=1,
        value="1. Sprints 1-4 ejecutados horizontal: cada sprint avanzo UNA capa en TODOS los MS")
ws.cell(row=31, column=1,
        value="2. Sprints 5-12 vertical por grupos: terminar grupo completo (backend+frontend+testing) "
              "antes del siguiente")
ws.cell(row=32, column=1,
        value="3. Grupo A (6 MS principales): Auth, Estudiantes, Instructores, Vehiculos, "
              "Asignaciones, Cobros")
ws.cell(row=33, column=1, value="4. Grupo B (2 MS secundarios): Notificaciones, Reportes")
ws.cell(row=34, column=1,
        value="5. Validacion continua + entregables funcionales por fase + JaCoCo 80% por PR")

# Bloque nuevo al final
next_row = 36
overview_extra = [
    "=== CAMBIO 2026-05-22: VERTICAL POR GRUPOS (Sprints 5-12) ===",
    "",
    "DIVISION DE MICROSERVICIOS EN GRUPOS:",
    "Grupo A - PRINCIPALES (6 MS): MS-Auth, MS-Estudiantes, MS-Instructores, MS-Vehiculos, "
    "MS-Asignaciones, MS-Cobros",
    "Grupo B - SECUNDARIOS (2 MS): MS-Notificaciones, MS-Reportes",
    "",
    "ROADMAP VIGENTE (Sprints 5-12) - detalle completo en hoja 'Sprints 5-12 Vigente' "
    "y en PLAN_FASES.md:",
    "  Sprint 5 (Fase 1 - Grupo A): Backend A pt.1 - CRUDs Auth + Estudiantes + Instructores + Vehiculos",
    "  Sprint 6 (Fase 1 - Grupo A): Backend A pt.2 - CRUDs Asignaciones (tripartita) + Cobros + Resilience4j",
    "  Sprint 7 (Fase 1 - Grupo A): Frontend completo Grupo A (vistas/forms/stores Pinia)",
    "  Sprint 8 (Fase 1 - Grupo A): Testing Grupo A (unit + integration + E2E Cypress 5 flujos)",
    "  Sprint 9 (Fase 2 - Grupo B): Backend Grupo B - Notificaciones (in-app + plantillas) "
    "+ Reportes (PDF/Excel)",
    "  Sprint 10 (Fase 2 - Grupo B): Frontend Grupo B (dashboard, reportes UI, notif dropdown)",
    "  Sprint 11 (Fase 2 - Grupo B): Testing Grupo B (unit + integration + E2E Cypress 3 flujos)",
    "  Sprint 12 (Fase 3 - Cierre): E2E cruzado + Performance JMeter + OWASP review + Deploy + Demo + Docs",
    "",
    "REFERENCIA: PLAN_FASES.md (fuente de verdad) - DECISIONES.md seccion 23 (ADR del cambio)",
]
for line in overview_extra:
    ws.cell(row=next_row, column=1, value=line)
    next_row += 1

print(f"Overview actualizado hasta fila {next_row}")

# === 2) Sprints Detallado: columna Estado ===
ws2 = wb["Sprints Detallado"]
ws2.cell(row=1, column=8, value="Estado")
ws2.cell(row=1, column=8).font = Font(bold=True)

fill_grey = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

for row in range(2, 179):
    ws2.cell(row=row, column=8, value="VIGENTE (ejecutado)")

for row in range(179, ws2.max_row + 1):
    ws2.cell(row=row, column=8,
             value="SUSTITUIDO - ver hoja 'Sprints 5-12 Vigente' y PLAN_FASES.md")
    for col in range(1, 9):
        ws2.cell(row=row, column=col).fill = fill_grey

print(f"Sprints Detallado: columna Estado agregada. {ws2.max_row} filas marcadas.")

# === 3) Hoja nueva 'Sprints 5-12 Vigente' ===
if "Sprints 5-12 Vigente" in wb.sheetnames:
    del wb["Sprints 5-12 Vigente"]
ws3 = wb.create_sheet("Sprints 5-12 Vigente", index=2)

headers = ["Sprint", "Fase", "Tarea", "MS", "Descripcion",
           "Criterios de Aceptacion", "Branch sugerido"]
for i, h in enumerate(headers, 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

data = [
    # Sprint 5
    (5, "Fase 1 - Grupo A", "T5.1", "MS-Auth",
     "Cierre modulo Configuracion + CRUD usuarios/roles. Endpoints GET/PUT /configuracion, "
     "CRUD tipos_curso, conceptos_facturacion, categorias_licencia, plantillas_email. "
     "CRUD /usuarios con asignacion de roles.",
     "Endpoints funcionando, validacion por @PreAuthorize, tests unit >=80%, "
     "JaCoCo verde, CI verde, PR mergeado.",
     "feature/sprint-5-1-auth-configuracion"),
    (5, "Fase 1 - Grupo A", "T5.2", "MS-Estudiantes",
     "CRUD completo: POST/GET/PUT/DELETE /estudiantes + sub-recursos /documentos, "
     "/contactos-emergencia, /progreso, /asistencia. Validacion cedula Ecuador. Soft delete.",
     "CRUD funcional, validaciones Ecuador OK, autorizacion por rol, soft delete, "
     "tests >=80%, JaCoCo verde.",
     "feature/sprint-5-2-estudiantes-crud"),
    (5, "Fase 1 - Grupo A", "T5.3", "MS-Instructores",
     "CRUD completo: /instructores + /certificaciones, /disponibilidad, /horarios-trabajo. "
     "Endpoint GET /instructores/{id}/disponibilidad?fecha. Validacion licencia profesional.",
     "CRUD funcional, endpoint disponibilidad OK, tests >=80%, JaCoCo verde.",
     "feature/sprint-5-3-instructores-crud"),
    (5, "Fase 1 - Grupo A", "T5.4", "MS-Vehiculos",
     "CRUD completo: /vehiculos + /mantenimientos, /combustible, /inspecciones, /documentos. "
     "Validacion placa Ecuador. Alerta SOAT (campo fecha_vencimiento_soat).",
     "CRUD funcional, validacion placa OK, alertas SOAT, tests >=80%, JaCoCo verde.",
     "feature/sprint-5-4-vehiculos-crud"),
    (5, "Fase 1 - Grupo A", "T5.5", "Cross-MS",
     "Evento UsuarioCreadoEvent propagado a MS-Estudiantes/Instructores: al crear "
     "estudiante/instructor, MS-Auth crea usuario y publica evento; consumers actualizan "
     "usuario_id en su entidad.",
     "Crear estudiante crea usuario automatico, idempotencia OK, evento procesado "
     "en <2s, tests integration OK.",
     "feature/sprint-5-5-evento-usuario-creado"),

    # Sprint 6
    (6, "Fase 1 - Grupo A", "T6.1", "MS-Asignaciones",
     "CRUD asignacion tripartita: POST /asignaciones valida disponibilidad de instructor "
     "(Feign), estudiante activo, vehiculo disponible. Deteccion de conflictos "
     "(mismo instructor/vehiculo mismo horario).",
     "Asignacion tripartita funcional, validaciones Feign OK, conflictos detectados, "
     "tests >=80%, JaCoCo verde.",
     "feature/sprint-6-1-asignaciones-crud"),
    (6, "Fase 1 - Grupo A", "T6.2", "MS-Asignaciones",
     "Reprogramacion + historial: PUT /asignaciones/{id}/reprogramar, tabla historial_estados. "
     "Eventos asignacion.creada, asignacion.reprogramada, asignacion.cancelada.",
     "Reprogramacion funcional, historial registrado, eventos publicados a RabbitMQ, tests OK.",
     "feature/sprint-6-2-asignaciones-reprogramar"),
    (6, "Fase 1 - Grupo A", "T6.3", "MS-Cobros",
     "CRUD facturas + pagos: POST/GET /facturas (con lineas), POST /pagos (soporta parciales, "
     "suma hasta total). Calculo saldo pendiente. Validacion RUC. "
     "Estados PENDIENTE/PAGADA/PARCIAL/ANULADA.",
     "Pagos parciales OK, saldo calculado correctamente, validacion RUC, "
     "tests >=80%, JaCoCo verde.",
     "feature/sprint-6-3-cobros-facturas-pagos"),
    (6, "Fase 1 - Grupo A", "T6.4", "MS-Cobros",
     "Reconciliacion + reporte de saldo: GET /cobros/estudiante/{id} devuelve estado de cuenta "
     "(facturas + pagos + saldo). Eventos pago.registrado, factura.emitida.",
     "Estado de cuenta OK, eventos publicados, tests OK.",
     "feature/sprint-6-4-cobros-reconciliacion"),
    (6, "Fase 1 - Grupo A", "T6.5", "Cross-MS",
     "Circuit Breakers Resilience4j en Feign clients de MS-Asignaciones a "
     "Estudiantes/Instructores/Vehiculos. Retry + fallback.",
     "Circuit breaker activa con MS caido, fallback devuelve 503 amigable, tests de fallo OK.",
     "feature/sprint-6-5-resilience4j"),

    # Sprint 7
    (7, "Fase 1 - Grupo A", "T7.1", "Frontend",
     "Setup base Vue 3 + Vite + TS + PrimeVue + Pinia + Router + Axios + interceptors JWT "
     "(cookie auto), MainLayout/AuthLayout, sidebar, header, breadcrumbs.",
     "App levanta con npm run dev, login funcional contra Gateway, layouts responsive, "
     "lint+type-check verde.",
     "feature/sprint-7-1-frontend-base"),
    (7, "Fase 1 - Grupo A", "T7.2", "Frontend Auth/Config",
     "LoginView, ForgotPasswordView, ResetPasswordView, ConfiguracionView (admin), "
     "UsuariosCRUDView, RolesView. Store useAuthStore.",
     "Login E2E OK, configuracion editable, CRUD usuarios funcional, vitest >=80%.",
     "feature/sprint-7-2-auth-config-ui"),
    (7, "Fase 1 - Grupo A", "T7.3", "Frontend Estudiantes",
     "ListaEstudiantesView (datatable + filtros), EstudianteFormView, EstudianteDetailView "
     "(tabs: datos, documentos, contactos, progreso, asistencia). Store useEstudiantesStore.",
     "CRUD UI funcional, validacion cedula en form, tabs cargan datos correctos, vitest >=80%.",
     "feature/sprint-7-3-estudiantes-ui"),
    (7, "Fase 1 - Grupo A", "T7.4", "Frontend Instructores",
     "ListaInstructoresView, InstructorFormView, InstructorDetailView (tabs: datos, "
     "certificaciones, disponibilidad, horarios). Calendario disponibilidad FullCalendar.",
     "CRUD UI funcional, calendario muestra disponibilidad, vitest >=80%.",
     "feature/sprint-7-4-instructores-ui"),
    (7, "Fase 1 - Grupo A", "T7.5", "Frontend Vehiculos",
     "ListaVehiculosView, VehiculoFormView, VehiculoDetailView (tabs: datos, mantenimientos, "
     "combustible, inspecciones, docs). Alertas SOAT en dashboard.",
     "CRUD UI funcional, validacion placa, alertas SOAT visibles, vitest >=80%.",
     "feature/sprint-7-5-vehiculos-ui"),
    (7, "Fase 1 - Grupo A", "T7.6", "Frontend Asignaciones",
     "CalendarioAsignacionesView (FullCalendar drag&drop), AsignacionFormView "
     "(wizard tripartita), AsignacionDetailView.",
     "Calendario funcional, wizard valida en vivo, drag&drop reprograma, vitest >=80%.",
     "feature/sprint-7-6-asignaciones-ui"),
    (7, "Fase 1 - Grupo A", "T7.7", "Frontend Cobros",
     "EstadoCuentaView (por estudiante), FacturaFormView, PagoFormView (soporta parciales), "
     "HistoricoCobrosView. Store useCobrosStore.",
     "Estado cuenta OK, pago parcial funcional, vitest >=80%.",
     "feature/sprint-7-7-cobros-ui"),

    # Sprint 8
    (8, "Fase 1 - Grupo A", "T8.1", "Backend (6 MS)",
     "Unit tests backend >=80% en los 6 MS del Grupo A. Subir umbral JaCoCo.",
     "JaCoCo verde >=80% en los 6 MS, CI verde.",
     "feature/sprint-8-1-backend-coverage"),
    (8, "Fase 1 - Grupo A", "T8.2", "Backend IT",
     "Integration tests con Testcontainers (Postgres + RabbitMQ) por MS. "
     "Tests *IT.java. Minimo 5 IT por MS.",
     "30+ IT pasando, RabbitMQ + Postgres reales OK.",
     "feature/sprint-8-2-backend-integration"),
    (8, "Fase 1 - Grupo A", "T8.3", "Frontend Tests",
     "Vitest >=80% en components, stores, composables del Grupo A.",
     "vitest coverage >=80%, lint+type-check verde.",
     "feature/sprint-8-3-frontend-coverage"),
    (8, "Fase 1 - Grupo A", "T8.4", "E2E Cypress",
     "5 flujos: 1) Login/logout, 2) Matricular estudiante, 3) Crear asignacion tripartita, "
     "4) Pago parcial, 5) Editar configuracion.",
     "5 flujos E2E pasan en CI sin flakiness.",
     "feature/sprint-8-4-e2e-cypress"),
    (8, "Fase 1 - Grupo A", "T8.5", "Bugfixes",
     "Slot para correcciones encontradas durante testing.",
     "Bugs registrados y resueltos. Documentar en docs/fase-1-cierre.md.",
     "feature/sprint-8-5-bugfixes"),

    # Sprint 9
    (9, "Fase 2 - Grupo B", "T9.1", "MS-Notificaciones",
     "Plantillas configurables CRUD (plantillas_email) + variables sustituibles. "
     "Tabla log_envios. Endpoint POST /notificaciones/test.",
     "Plantillas editables, log de envios poblado, test endpoint OK, tests >=80%.",
     "feature/sprint-9-1-notif-plantillas"),
    (9, "Fase 2 - Grupo B", "T9.2", "MS-Notificaciones",
     "Notificaciones in-app: tabla notificaciones, GET /notificaciones?leidas=false, "
     "PATCH /notificaciones/{id}/leer. Consumer adicional para eventos Grupo A.",
     "Polling devuelve notif en <30s tras evento, marcar leida funciona, tests OK.",
     "feature/sprint-9-2-notif-in-app"),
    (9, "Fase 2 - Grupo B", "T9.3", "MS-Reportes",
     "Endpoints reportes operativos: estudiantes activos, instructores con horas, "
     "vehiculos SOAT por vencer, asistencia por curso. Queries cross-MS via Feign.",
     "5+ reportes operativos funcionando, datos correctos cross-MS, tests OK.",
     "feature/sprint-9-3-reportes-operativos"),
    (9, "Fase 2 - Grupo B", "T9.4", "MS-Reportes",
     "Endpoints reportes financieros: ingresos por periodo, saldos pendientes, "
     "morosidad, recibos emitidos. KPIs para dashboard.",
     "4+ reportes financieros funcionando, KPIs correctos, tests OK.",
     "feature/sprint-9-4-reportes-financieros"),
    (9, "Fase 2 - Grupo B", "T9.5", "MS-Reportes",
     "Exportacion PDF (Thymeleaf + OpenPDF) + Excel (Apache POI). "
     "Endpoint POST /reportes/exportar?formato=pdf|excel.",
     "Export PDF y Excel funcionando, ambos con datos correctos.",
     "feature/sprint-9-5-reportes-export"),
    (9, "Fase 2 - Grupo B", "T9.6", "MS-Reportes",
     "Cache de reportes (Caffeine) con TTL configurable. Tabla ejecuciones_reporte para auditoria.",
     "Cache hit en segunda llamada, TTL respetado, tabla poblada.",
     "feature/sprint-9-6-reportes-cache"),

    # Sprint 10
    (10, "Fase 2 - Grupo B", "T10.1", "Frontend Notif",
     "<NotificacionesDropdown /> en header con badge count no leidas. Polling cada 30s. "
     "Marcar como leida. Store useNotificacionesStore.",
     "Dropdown muestra notif, polling activo, marcar leida OK, vitest >=80%.",
     "feature/sprint-10-1-notif-dropdown"),
    (10, "Fase 2 - Grupo B", "T10.2", "Frontend Notif Config",
     "PlantillasEmailView (CRUD + preview), LogEnviosView (historico, filtros).",
     "Plantillas editables con preview, log filtrable, vitest >=80%.",
     "feature/sprint-10-2-notif-config-ui"),
    (10, "Fase 2 - Grupo B", "T10.3", "Frontend Dashboard",
     "DashboardView con KPIs (Chart.js): estudiantes activos, ingresos mes, clases hoy, "
     "vehiculos SOAT por vencer. Cards interactivas.",
     "Dashboard carga en <2s, KPIs actualizados, navegacion desde cards OK.",
     "feature/sprint-10-3-dashboard"),
    (10, "Fase 2 - Grupo B", "T10.4", "Frontend Reportes Op.",
     "ReporteEstudiantesView, ReporteInstructoresView, ReporteVehiculosView, "
     "ReporteAsistenciaView. Filtros, tabla, export PDF/Excel.",
     "4 reportes operativos UI funcionales, export PDF/Excel desde UI.",
     "feature/sprint-10-4-reportes-operativos-ui"),
    (10, "Fase 2 - Grupo B", "T10.5", "Frontend Reportes Fin.",
     "ReporteIngresosView, ReporteMorosidadView, ReporteRecibosView. "
     "Charts (barras, lineas). Export PDF/Excel.",
     "3 reportes financieros UI funcionales con charts, export OK.",
     "feature/sprint-10-5-reportes-financieros-ui"),

    # Sprint 11
    (11, "Fase 2 - Grupo B", "T11.1", "Backend (2 MS)",
     "Unit tests >=80% en MS-Notificaciones y MS-Reportes.",
     "JaCoCo verde >=80% en los 2 MS, CI verde.",
     "feature/sprint-11-1-backend-coverage"),
    (11, "Fase 2 - Grupo B", "T11.2", "Backend IT",
     "Testcontainers para MS-Notif (RabbitMQ+Postgres+GreenMail SMTP), "
     "MS-Reportes (Postgres+Feign mocks).",
     "IT pasando, emails reales capturados en GreenMail, Feign mockeado OK.",
     "feature/sprint-11-2-backend-integration"),
    (11, "Fase 2 - Grupo B", "T11.3", "Frontend Tests",
     "Vitest >=80% en Dashboard, Reportes, Notificaciones.",
     "vitest coverage >=80%.",
     "feature/sprint-11-3-frontend-coverage"),
    (11, "Fase 2 - Grupo B", "T11.4", "E2E Cypress",
     "3 flujos: 1) Evento->email->notif in-app->leida, 2) Generar reporte op + export PDF, "
     "3) Editar plantilla + envio prueba.",
     "3 flujos E2E pasan en CI.",
     "feature/sprint-11-4-e2e-cypress"),
    (11, "Fase 2 - Grupo B", "T11.5", "Bugfixes",
     "Slot para correcciones.",
     "Bugs registrados/resueltos. Documentar en docs/fase-2-cierre.md.",
     "feature/sprint-11-5-bugfixes"),

    # Sprint 12
    (12, "Fase 3 - Cierre", "T12.1", "E2E cruzado",
     "Cypress: flujo completo estudiante. Matricula -> factura -> asignar 5 clases -> "
     "asistencia -> pago parcial -> pago restante -> recibo email -> reporte financiero.",
     "Flujo E2E cruzado pasa, datos consistentes cross-MS.",
     "feature/sprint-12-1-e2e-cruzado"),
    (12, "Fase 3 - Cierre", "T12.2", "Performance",
     "JMeter: 50 usuarios concurrentes. Target p95 <500ms. Identificar bottlenecks, "
     "agregar indices BD, ajustar pool HikariCP.",
     "p95 <500ms en endpoints criticos, reporte JMeter documentado.",
     "feature/sprint-12-2-performance"),
    (12, "Fase 3 - Cierre", "T12.3", "OWASP",
     "Checklist OWASP Top 10 completo. mvn dependency-check.",
     "Sin vulnerabilidades criticas, checklist documentado.",
     "feature/sprint-12-3-owasp"),
    (12, "Fase 3 - Cierre", "T12.4", "Rate Limiting",
     "100 req/min/IP en API Gateway (Bucket4j o filtro custom).",
     "Rate limit funciona, headers X-RateLimit-* presentes.",
     "feature/sprint-12-4-rate-limiting"),
    (12, "Fase 3 - Cierre", "T12.5", "Limpieza",
     "Eliminar TODOs, codigo muerto, console.log, deps no usadas. "
     "Scheduler cleanup refresh tokens expirados.",
     "Lint clean, scheduler activo.",
     "feature/sprint-12-5-limpieza"),
    (12, "Fase 3 - Cierre", "T12.6", "Documentacion final",
     "README maestro, docs/runbook.md, docs/manual-usuario.md, C4 diagrams, "
     "OpenAPI specs publicados.",
     "Docs completos y actualizados.",
     "feature/sprint-12-6-docs"),
    (12, "Fase 3 - Cierre", "T12.7", "Deploy",
     "Oracle Cloud Free Tier (fallback DigitalOcean $6). Nginx + Let's Encrypt. Backup diario.",
     "Sistema accesible publicamente, HTTPS, backup automatizado.",
     "feature/sprint-12-7-deploy"),
    (12, "Fase 3 - Cierre", "T12.8", "Demo + Cierre",
     "Video demo (15 min). Slides. Entrega final titulacion. Tag v1.0.0.",
     "Demo grabada, entrega aceptada, tag publicado.",
     "feature/sprint-12-8-demo"),
]

fase1_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
fase2_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
fase3_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)

for i, row_data in enumerate(data, 2):
    fase = row_data[1]
    if "Fase 1" in fase:
        fill = fase1_fill
    elif "Fase 2" in fase:
        fill = fase2_fill
    else:
        fill = fase3_fill
    for col_idx, val in enumerate(row_data, 1):
        c = ws3.cell(row=i, column=col_idx, value=val)
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = thin_border

ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 8
ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 70
ws3.column_dimensions["F"].width = 50
ws3.column_dimensions["G"].width = 40
ws3.row_dimensions[1].height = 30
ws3.freeze_panes = "A2"

print("Hoja nueva 'Sprints 5-12 Vigente' creada con " + str(len(data)) + " filas")

wb.save("SPRINTS_PLAN.xlsx")
print("\n[OK] SPRINTS_PLAN.xlsx actualizado correctamente")
print("Hojas finales:", wb.sheetnames)
